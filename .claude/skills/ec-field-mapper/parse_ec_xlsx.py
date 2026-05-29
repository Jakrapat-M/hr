#!/usr/bin/env python3
"""
EC Field Mapper — parse the BA "EC list of fields" xlsx into a normalized model,
diff it against what's implemented in src/, and emit scaffold code for gaps.

No JS/npm dependency: xlsx is read with openpyxl (python, offline). Output is
staged under .omc/ec-mapper/ for human review — this script never writes into src/.

Usage:
  python3 parse_ec_xlsx.py parse   --xlsx <file> [--out <dir>]
  python3 parse_ec_xlsx.py report  --xlsx <file> [--repo <root>] [--out <dir>]
  python3 parse_ec_xlsx.py scaffold --xlsx <file> --picklist <ID>   [--out <dir>]
  python3 parse_ec_xlsx.py scaffold --xlsx <file> --field "<UI Field>" --process Hiring [--out <dir>]

The "report" subcommand is the main entry: it parses, routes, diffs coverage,
and writes report.md + normalized.json into <out> (default .omc/ec-mapper/).
"""
from __future__ import annotations
import argparse, json, os, re, sys
from collections import defaultdict, Counter

CANON_SHEET = "Employee file "          # trailing space is intentional (real sheet name)
LOV_SHEET = "LOV"
SUMMARY_SHEET = "summary of hiring maintain"

# Canonical-sheet column indices (header row 4, data row 5+).
COL = dict(process=0, section=1, subSection=2, uiField=3, mandatory=4,
           employeeGroup=5, editable=11, default=12, validation=13,
           hrLogic=14, hrConfirm=15, remark=17, editType=18, maintainBy=19)

# BA Section -> hire-wizard step (process=Hiring). Fallback: the section name.
SECTION_TO_STEP = {
    "identity": "StepIdentity",
    "personal information": "StepBiographical",
    "job information": "StepJob",
    "compensation information": "StepJob",
    "contact": "StepContact",
    "contact information": "StepContact",
    "emergency contact": "StepContact",
}

def _norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip())

def _slug_field(s):
    """Normalize a UI Field label for fuzzy coverage matching."""
    s = _norm(s).lower()
    s = re.sub(r"\(.*?\)", "", s)          # drop "(EN)" / "(TH)" qualifiers
    s = re.sub(r"[^a-z0-9 ]", "", s)
    return _norm(s)

# ──────────────────────────────────────────────────────────────────────────────
def load_wb(path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl not available. Install it in a venv or use the "
                 "system python that has it (this skill assumes offline openpyxl).")
    if not os.path.exists(path):
        sys.exit(f"ERROR: xlsx not found: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

def parse_fields(wb):
    if CANON_SHEET not in wb.sheetnames:
        # tolerate a stripped/renamed sheet
        cand = [s for s in wb.sheetnames if s.strip().lower() == CANON_SHEET.strip().lower()]
        if not cand:
            sys.exit(f"ERROR: sheet '{CANON_SHEET}' not found. Sheets: {wb.sheetnames}")
        sheet = cand[0]
    else:
        sheet = CANON_SHEET
    ws = wb[sheet]
    fields = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        process = _norm(r[COL["process"]]) if len(r) > COL["process"] else ""
        ui = _norm(r[COL["uiField"]]) if len(r) > COL["uiField"] else ""
        if not process or not ui:
            continue
        def g(k):
            i = COL[k]
            return _norm(r[i]) if len(r) > i else ""
        fields.append(dict(
            process=process, section=g("section"), subSection=g("subSection"),
            uiField=ui, mandatory=g("mandatory"), employeeGroup=g("employeeGroup"),
            editable=g("editable"), default=g("default"), validation=g("validation"),
            remark=g("remark"), editType=g("editType"), maintainBy=g("maintainBy"),
            slug=_slug_field(ui),
        ))
    return fields

def parse_lov(wb):
    """Return {picklistId: [{code,labelEn,labelTh,parent,status}]}."""
    if LOV_SHEET not in wb.sheetnames:
        return {}
    ws = wb[LOV_SHEET]
    lov = defaultdict(list)
    # header is row 2 (1-indexed); data row 3+
    for r in ws.iter_rows(min_row=3, values_only=True):
        pid = _norm(r[0]) if len(r) > 0 else ""
        if not pid:
            continue
        lov[pid].append(dict(
            status=_norm(r[1]) if len(r) > 1 else "",
            code=_norm(r[2]) if len(r) > 2 else "",
            parent=_norm(r[4]) if len(r) > 4 else "",
            labelEn=_norm(r[6]) if len(r) > 6 else "",
            labelTh=_norm(r[7]) if len(r) > 7 else "",
        ))
    return dict(lov)

def route(field):
    """Decide the UI surface for a field."""
    sec = field["section"].lower()
    if field["process"].lower() == "hiring":
        return dict(kind="wizard-step", name=SECTION_TO_STEP.get(sec, field["section"] or "StepUnknown"))
    # maintain -> employee profile section (FieldCard group keyed by Section)
    return dict(kind="profile-section", name=field["section"] or "Profile")

def detect_lov_ref(field, lov_ids):
    """Heuristic: a field references a picklist if its validation/editable text
    contains an LOV Picklist ID token."""
    blob = f"{field['validation']} {field['editable']} {field['default']}".upper()
    for pid in lov_ids:
        if pid and pid.upper() in blob:
            return pid
    return None

# ── Coverage diff ──────────────────────────────────────────────────────────────
COVERAGE_CORPUS = [
    "src/frontend/src/components/admin/hire/steps",          # wizard steps (dir)
    "src/frontend/src/app/[locale]/admin/hire/steps",        # alt step location
    "src/frontend/src/lib/humi-mock-data.ts",                # profile job/personal arrays
    "src/frontend/src/lib/sta82-employee-profile-field-spec.ts",
    "src/frontend/src/lib/admin/hire/picklists",             # picklist registry (dir)
]

def _read_corpus(repo):
    text = []
    for rel in COVERAGE_CORPUS:
        p = os.path.join(repo, rel)
        if os.path.isdir(p):
            for root, _, files in os.walk(p):
                for f in files:
                    if f.endswith((".ts", ".tsx", ".json")):
                        try:
                            text.append(open(os.path.join(root, f), encoding="utf-8").read())
                        except Exception:
                            pass
        elif os.path.isfile(p):
            try:
                text.append(open(p, encoding="utf-8").read())
            except Exception:
                pass
    return "\n".join(text).lower()

def coverage(fields, repo):
    corpus = _read_corpus(repo)
    for f in fields:
        slug = f["slug"]
        # implemented if the normalized label (>=4 chars) appears in the corpus
        f["implemented"] = bool(slug) and len(slug) >= 4 and slug in corpus
    return fields

# ── Scaffold generators ─────────────────────────────────────────────────────────
def scaffold_picklist(pid, options):
    const = re.sub(r"[^A-Z0-9]+", "_", pid.upper()).strip("_") + "_OPTIONS"
    lines = [f"// {pid}.ts — generated by ec-field-mapper from BA LOV sheet. Review before use.",
             "import type { PicklistDefinition } from './types'", "",
             f"export const {const}: PicklistDefinition = ["]
    for o in options:
        if o["status"] and o["status"].upper() not in ("A", ""):
            continue  # active picklist values only
        code = o["code"].replace("'", "\\'")
        th = (o["labelTh"] or o["labelEn"]).replace("'", "\\'")
        en = (o["labelEn"] or o["labelTh"]).replace("'", "\\'")
        lines.append(f"  {{ id: '{code}', labelTh: '{th}', labelEn: '{en}' }},")
    lines.append("] as const")
    lines.append("")
    return const, "\n".join(lines)

def scaffold_jsx(field, lov_const=None):
    fid = re.sub(r"[^a-z0-9]+", "-", _slug_field(field["uiField"]))
    name = re.sub(r"[^a-zA-Z0-9]+", " ", field["uiField"]).title().replace(" ", "")
    name = name[0].lower() + name[1:] if name else "field"
    req = "Required" in field["mandatory"]
    label = field["uiField"]
    if lov_const:
        return (f"<fieldset>\n"
                f"  <label htmlFor=\"{fid}\" className=\"humi-label\">{label}{' *' if req else ''}</label>\n"
                f"  <select id=\"{fid}\" className=\"humi-input w-full\" value={{form.{name}}} "
                f"onChange={{e => set('{name}', e.target.value)}}>\n"
                f"    <option value=\"\">—</option>\n"
                f"    {{{lov_const}.map(o => (\n"
                f"      <option key={{o.id}} value={{o.id}}>{{pickLabel(o, locale)}}</option>\n"
                f"    ))}}\n"
                f"  </select>\n"
                f"</fieldset>")
    itype = "date" if re.search(r"date", label, re.I) else "text"
    return (f"<fieldset>\n"
            f"  <label htmlFor=\"{fid}\" className=\"humi-label\">{label}{' *' if req else ''}</label>\n"
            f"  <input id=\"{fid}\" type=\"{itype}\" className=\"humi-input w-full\" value={{form.{name}}} "
            f"onChange={{e => set('{name}', e.target.value)}} />\n"
            f"</fieldset>")

def scaffold_i18n(field):
    key = re.sub(r"[^a-zA-Z0-9]+", "_", _slug_field(field["uiField"])).strip("_")
    return key, field["uiField"]  # en label; th to be supplied by translator

# ── Subcommands ──────────────────────────────────────────────────────────────────
def cmd_parse(args):
    wb = load_wb(args.xlsx)
    fields = parse_fields(wb)
    lov = parse_lov(wb)
    out = dict(
        source=os.path.basename(args.xlsx),
        counts=dict(fields=len(fields),
                    byProcess=dict(Counter(f["process"] for f in fields)),
                    picklists=len(lov)),
        fields=fields,
        picklistIds=sorted(lov.keys()),
    )
    _emit_json(args.out, "normalized.json", out)
    print(f"parsed {len(fields)} fields, {len(lov)} picklists -> {args.out}/normalized.json")

def cmd_report(args):
    wb = load_wb(args.xlsx)
    fields = parse_fields(wb)
    lov = parse_lov(wb)
    lov_ids = list(lov.keys())
    for f in fields:
        f["surface"] = route(f)
        f["lovRef"] = detect_lov_ref(f, lov_ids)
    coverage(fields, args.repo)

    impl = [f for f in fields if f["implemented"]]
    missing = [f for f in fields if not f["implemented"]]
    by_surface = defaultdict(list)
    for f in fields:
        by_surface[f["surface"]["name"]].append(f)

    md = []
    md.append(f"# EC Field Mapper Report\n")
    md.append(f"- Source: `{os.path.basename(args.xlsx)}`")
    md.append(f"- Total fields: **{len(fields)}** "
              f"(Hiring {sum(1 for f in fields if f['process'].lower()=='hiring')}, "
              f"maintain {sum(1 for f in fields if f['process'].lower()=='maintain')})")
    md.append(f"- Picklists in LOV sheet: **{len(lov)}**")
    md.append(f"- Coverage: **{len(impl)} implemented / {len(missing)} missing** "
              f"({100*len(impl)//max(1,len(fields))}%)\n")

    md.append("## Coverage by surface\n")
    md.append("| Surface | Total | Implemented | Missing |")
    md.append("|---|---|---|---|")
    for name in sorted(by_surface):
        g = by_surface[name]
        md.append(f"| {name} | {len(g)} | {sum(1 for f in g if f['implemented'])} "
                  f"| {sum(1 for f in g if not f['implemented'])} |")

    md.append("\n## Missing fields (scaffold candidates)\n")
    md.append("| Process | Section | Sub-section | UI Field | Mandatory | LOV ref | Surface |")
    md.append("|---|---|---|---|---|---|---|")
    for f in missing[:400]:
        md.append(f"| {f['process']} | {f['section']} | {f['subSection']} | {f['uiField']} "
                  f"| {f['mandatory']} | {f['lovRef'] or ''} | {f['surface']['name']} |")
    if len(missing) > 400:
        md.append(f"\n_(+{len(missing)-400} more — see normalized.json)_")

    _emit_json(args.out, "normalized.json",
               dict(fields=fields, picklistIds=sorted(lov.keys())))
    _emit_text(args.out, "report.md", "\n".join(md))
    print(f"report -> {args.out}/report.md  ({len(impl)} impl / {len(missing)} missing)")

def cmd_scaffold(args):
    wb = load_wb(args.xlsx)
    lov = parse_lov(wb)
    sdir = os.path.join(args.out, "scaffold")
    os.makedirs(sdir, exist_ok=True)
    if args.picklist:
        if args.picklist not in lov:
            sys.exit(f"ERROR: picklist '{args.picklist}' not in LOV sheet")
        const, body = scaffold_picklist(args.picklist, lov[args.picklist])
        fname = re.sub(r"[^a-zA-Z0-9]+", "", args.picklist.title()) + ".ts"
        _emit_text(sdir, fname, body)
        print(f"scaffold picklist {const} -> {sdir}/{fname}")
    elif args.field:
        fields = parse_fields(wb)
        match = [f for f in fields if _slug_field(f["uiField"]) == _slug_field(args.field)
                 and (not args.process or f["process"].lower() == args.process.lower())]
        if not match:
            sys.exit(f"ERROR: field '{args.field}' not found")
        f = match[0]
        ref = detect_lov_ref(f, list(lov.keys()))
        const = None
        if ref:
            const, body = scaffold_picklist(ref, lov[ref])
            _emit_text(sdir, re.sub(r"[^a-zA-Z0-9]+", "", ref.title()) + ".ts", body)
        jsx = scaffold_jsx(f, const)
        key, en = scaffold_i18n(f)
        out = (f"// JSX (place in {f['surface']['name'] if 'surface' in f else route(f)['name']})\n{jsx}\n\n"
               f"// i18n — add to messages/en.json and messages/th.json:\n"
               f'//   en: "{key}": "{en}"\n//   th: "{key}": "<TH translation>"\n')
        _emit_text(sdir, key + ".scaffold.txt", out)
        print(f"scaffold field '{f['uiField']}' -> {sdir}/{key}.scaffold.txt")
    else:
        sys.exit("ERROR: scaffold needs --picklist <ID> or --field <name>")

def _emit_json(out, name, obj):
    os.makedirs(out, exist_ok=True)
    with open(os.path.join(out, name), "w", encoding="utf-8") as fh:
        json.dump(obj, fh, ensure_ascii=False, indent=2)

def _emit_text(out, name, text):
    os.makedirs(out, exist_ok=True)
    with open(os.path.join(out, name), "w", encoding="utf-8") as fh:
        fh.write(text)

def main():
    ap = argparse.ArgumentParser(description="EC field mapper (BA xlsx -> mapping/coverage/scaffold)")
    sub = ap.add_subparsers(dest="cmd", required=True)
    for cmd in ("parse", "report", "scaffold"):
        p = sub.add_parser(cmd)
        p.add_argument("--xlsx", required=True)
        p.add_argument("--out", default=".omc/ec-mapper")
        p.add_argument("--repo", default=".")
        if cmd == "scaffold":
            p.add_argument("--picklist")
            p.add_argument("--field")
            p.add_argument("--process")
    args = ap.parse_args()
    {"parse": cmd_parse, "report": cmd_report, "scaffold": cmd_scaffold}[args.cmd](args)

if __name__ == "__main__":
    main()
